#!/usr/bin/env python3
import os, sys
output_dir = sys.argv[1] if len(sys.argv) > 1 else "."
os.makedirs(output_dir, exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Cores ──────────────────────────────────────────────────
COR_HEADER    = "1A3C5E"   # azul IDAF
COR_OBG       = "FFF2CC"   # amarelo — campo obrigatório
COR_OPC       = "EBF3FB"   # azul claro — campo opcional
COR_EXEMPLO   = "F2F2F2"   # cinza — linha de exemplo
COR_BRANCO    = "FFFFFF"

def borda_fina():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def celula_header(ws, row, col, texto, largura=None):
    c = ws.cell(row=row, column=col, value=texto)
    c.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill      = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = borda_fina()
    if largura:
        ws.column_dimensions[get_column_letter(col)].width = largura
    return c

def celula_sub(ws, row, col, texto, cor=COR_OBG):
    c = ws.cell(row=row, column=col, value=texto)
    c.font      = Font(bold=False, size=9, name="Arial", italic=True, color="555555")
    c.fill      = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = borda_fina()
    return c

def celula_exemplo(ws, row, col, texto):
    c = ws.cell(row=row, column=col, value=texto)
    c.font      = Font(size=9, name="Arial", color="333333")
    c.fill      = PatternFill("solid", fgColor=COR_EXEMPLO)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = borda_fina()
    return c

def celula_vazia(ws, row, col):
    c = ws.cell(row=row, column=col, value=None)
    c.fill      = PatternFill("solid", fgColor=COR_BRANCO)
    c.border    = borda_fina()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

# ════════════════════════════════════════════════════════════
# TEMPLATE CONTRATOS
# ════════════════════════════════════════════════════════════
wb = Workbook()

# ── Aba Contratos ──────────────────────────────────────────
ws = wb.active
ws.title = "Contratos"
ws.freeze_panes = "A4"
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 40

# Título
ws.merge_cells("A1:N1")
t = ws["A1"]
t.value     = "IDAF/AC — Modelo de Importação de Contratos"
t.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
t.fill      = PatternFill("solid", fgColor=COR_HEADER)
t.alignment = Alignment(horizontal="center", vertical="center")

# Legenda
ws.merge_cells("A2:N2")
leg = ws["A2"]
leg.value     = "🟡 Fundo amarelo = OBRIGATÓRIO   🔵 Fundo azul = OPCIONAL   Datas no formato DD/MM/AAAA   Valores em reais (ex: 15000,00)"
leg.font      = Font(size=9, name="Arial", color="333333")
leg.fill      = PatternFill("solid", fgColor="FFFACD")
leg.alignment = Alignment(horizontal="center", vertical="center")

# Colunas: (titulo, subtitulo, largura, obrigatorio, exemplo)
colunas = [
    ("numero",             "Nº do Contrato",          18, True,  "089/2023"),
    ("objeto",             "Descrição do Objeto",      45, True,  "Contratação de empresa para serviços de..."),
    ("fornecedor",         "Razão Social",             35, False, "EMPRESA EXEMPLO LTDA"),
    ("cnpj",               "CNPJ (XX.XXX.XXX/XXXX-XX)",20, False, "12.345.678/0001-90"),
    ("valor_total",        "Valor Total (R$)",         18, True,  "150000,00"),
    ("saldo_atual",        "Saldo Atual (R$)",         18, False, "150000,00"),
    ("data_assinatura",    "Data de Assinatura",       18, False, "15/01/2023"),
    ("data_inicio",        "Início da Vigência",       18, False, "15/01/2023"),
    ("data_vencimento",    "Data de Vencimento",       18, True,  "14/01/2024"),
    ("numero_processo",    "Nº do Processo SEI",       22, False, "0052.013538.00005/2024-50"),
    ("modalidade",         "Modalidade",               22, False, "pregao_eletronico"),
    ("gestor_nome",        "Nome do Gestor",           25, False, "João da Silva"),
    ("fiscal_nome",        "Nome do Fiscal",           25, False, "Maria Souza"),
    ("observacoes",        "Observações",              30, False, "Contrato vigente"),
]

for i, (nome, titulo, largura, obrig, exemplo) in enumerate(colunas, start=1):
    cor = COR_OBG if obrig else COR_OPC
    celula_header(ws, 3, i, titulo.upper(), largura)
    celula_sub(ws, 4, i, f"({nome})", cor)
    celula_exemplo(ws, 5, i, exemplo)
    for row in range(6, 106):
        celula_vazia(ws, row, i)

ws.row_dimensions[4].height = 22
ws.row_dimensions[5].height = 20

# ── Aba Modalidades (referência) ───────────────────────────
ws2 = wb.create_sheet("Modalidades (referência)")
ref = [
    ("Valor a usar na planilha", "Descrição"),
    ("pregao_eletronico",        "Pregão Eletrônico"),
    ("pregao_presencial",        "Pregão Presencial"),
    ("concorrencia",             "Concorrência"),
    ("tomada_de_precos",         "Tomada de Preços"),
    ("convite",                  "Convite"),
    ("dispensa",                 "Dispensa de Licitação"),
    ("inexigibilidade",          "Inexigibilidade"),
    ("chamamento_publico",       "Chamamento Público"),
]
for i, (val, desc) in enumerate(ref, start=1):
    bold = i == 1
    c1 = ws2.cell(row=i, column=1, value=val)
    c2 = ws2.cell(row=i, column=2, value=desc)
    for c in (c1, c2):
        c.font = Font(bold=bold, name="Arial", size=10,
                      color="FFFFFF" if bold else "000000")
        c.fill = PatternFill("solid", fgColor=COR_HEADER if bold else COR_BRANCO)
        c.border = borda_fina()
        c.alignment = Alignment(vertical="center")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 30

wb.save(os.path.join(output_dir, "template_contratos.xlsx"))
print("template_contratos.xlsx gerado")

# ════════════════════════════════════════════════════════════
# TEMPLATE LICITAÇÕES
# ════════════════════════════════════════════════════════════
wb2 = Workbook()
ws3 = wb2.active
ws3.title = "Licitações"
ws3.freeze_panes = "A4"
ws3.row_dimensions[1].height = 36
ws3.row_dimensions[2].height = 28
ws3.row_dimensions[3].height = 40

ws3.merge_cells("A1:J1")
t2 = ws3["A1"]
t2.value     = "IDAF/AC — Modelo de Importação de Licitações"
t2.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
t2.fill      = PatternFill("solid", fgColor=COR_HEADER)
t2.alignment = Alignment(horizontal="center", vertical="center")

ws3.merge_cells("A2:J2")
leg2 = ws3["A2"]
leg2.value     = "🟡 Fundo amarelo = OBRIGATÓRIO   🔵 Fundo azul = OPCIONAL   Datas no formato DD/MM/AAAA   Valores em reais (ex: 50000,00)"
leg2.font      = Font(size=9, name="Arial", color="333333")
leg2.fill      = PatternFill("solid", fgColor="FFFACD")
leg2.alignment = Alignment(horizontal="center", vertical="center")

colunas_lic = [
    ("numero_processo",          "Nº do Processo SEI",        28, True,  "0052.013537.00031/2023-06"),
    ("objeto",                   "Descrição do Objeto",       45, True,  "Contratação de empresa para..."),
    ("modalidade",               "Modalidade",                22, False, "pregao_eletronico"),
    ("status",                   "Status",                    22, False, "em_andamento"),
    ("valor_estimado",           "Valor Estimado (R$)",       20, False, "50000,00"),
    ("data_abertura",            "Data de Abertura",          18, False, "04/12/2023"),
    ("data_prevista_conclusao",  "Previsão de Conclusão",     20, False, "30/01/2024"),
    ("responsavel",              "Pregoeiro/Responsável",     25, False, "Sandra Maria Nunes Barbosa"),
    ("numero_licitacao",         "Nº da Licitação",           18, False, "099/2023"),
    ("observacoes",              "Observações",               30, False, "Licitação em andamento"),
]

for i, (nome, titulo, largura, obrig, exemplo) in enumerate(colunas_lic, start=1):
    cor = COR_OBG if obrig else COR_OPC
    celula_header(ws3, 3, i, titulo.upper(), largura)
    celula_sub(ws3, 4, i, f"({nome})", cor)
    celula_exemplo(ws3, 5, i, exemplo)
    for row in range(6, 106):
        celula_vazia(ws3, row, i)

ws3.row_dimensions[4].height = 22
ws3.row_dimensions[5].height = 20

# Aba status referência
ws4 = wb2.create_sheet("Status (referência)")
status_ref = [
    ("Valor a usar",              "Descrição"),
    ("em_andamento",              "Em Andamento"),
    ("aguardando_homologacao",    "Aguardando Homologação"),
    ("homologada",                "Homologada"),
    ("deserta",                   "Deserta"),
    ("fracassada",                "Fracassada"),
    ("cancelada",                 "Cancelada"),
    ("suspensa",                  "Suspensa"),
]
for i, (val, desc) in enumerate(status_ref, start=1):
    bold = i == 1
    c1 = ws4.cell(row=i, column=1, value=val)
    c2 = ws4.cell(row=i, column=2, value=desc)
    for c in (c1, c2):
        c.font = Font(bold=bold, name="Arial", size=10,
                      color="FFFFFF" if bold else "000000")
        c.fill = PatternFill("solid", fgColor=COR_HEADER if bold else COR_BRANCO)
        c.border = borda_fina()
        c.alignment = Alignment(vertical="center")
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 30

wb2.save(os.path.join(output_dir, "template_licitacoes.xlsx"))
print("template_licitacoes.xlsx gerado")
